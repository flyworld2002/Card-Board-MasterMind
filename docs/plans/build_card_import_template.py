"""
One-off script that generates docs/plans/card_import_template.xlsx —
a sample spreadsheet for Fei to fill out for the upcoming Excel-to-staging
importer (docs/plans/listing-pricing-system.md, PLANNING session 14, item 2).

Not wired into main.py — this only builds the template file itself, not
the importer that will read it.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "card_name", "set_name", "set_code", "card_number", "rarity",
    "foil_type", "foil_pattern", "texture", "material", "size",
    "stamp_type", "source_type", "is_promo", "is_first_edition",
    "image_url",
    "condition", "quantity", "price", "purchase_date", "source",
    "reference_id", "notes",
]

EXAMPLES = [
    # Plain existing card, no variant axes, no promo flags.
    ["Charmeleon", "Vivid Voltage", "swsh4", "24", "Uncommon",
     "", "", "", "", "",
     "", "", "", "",
     "",
     "Near Mint", 4, 0.35, "2026-07-15", "local_shop",
     "", ""],
    # Reverse holo variant of an existing card.
    ["Pikachu", "Vivid Voltage", "swsh4", "50", "Common",
     "reverse_holo", "", "", "", "",
     "", "", "", "",
     "",
     "Near Mint", 2, 1.20, "2026-07-15", "card_show",
     "", ""],
    # New/uncatalogued promo card — no existing card_master match expected,
    # so this row is meant to demonstrate the manual-create fallback path.
    ["Charizard ex", "Scarlet & Violet Promos", "svp", "199", "Promo",
     "holo", "", "", "", "jumbo",
     "pokemon_center", "", True, False,
     "",
     "Near Mint", 1, 45.00, "2026-07-20", "ebay",
     "PROMO-SVP-199", "Pokemon Center exclusive, sealed"],
]

FIELD_GUIDE = [
    ("Column", "Required?", "Notes"),
    ("card_name", "Required", "Exact card name."),
    ("set_name", "Required", "Must match an existing set's name to reuse it; a new name creates a new set (needs set_code too)."),
    ("set_code", "Required if set is new", "e.g. swsh4, svp. Leave blank if set_name already exists in the catalog."),
    ("card_number", "Required", "As printed on the card, e.g. 24, 199."),
    ("rarity", "Optional", "e.g. Common, Uncommon, Rare, Double Rare, Promo."),
    ("foil_type", "Optional", "non_holo, holo, reverse_holo — leave blank for a plain normal card."),
    ("foil_pattern", "Optional", "poke_ball, master_ball, friend_ball, love_ball, quick_ball, dusk_ball, team_rocket, energy_symbol."),
    ("texture", "Optional", "cosmos, hd_cosmos, galaxy_cosmos."),
    ("material", "Optional", "metal."),
    ("size", "Optional", "jumbo."),
    ("stamp_type", "Optional", "1st_edition, pokemon_center, prerelease, pokemon_day, mega_evolution, prismatic_evolution."),
    ("source_type", "Optional", "deck_exclusive, product_exclusive, box_topper, stamp_promo."),
    ("is_promo", "Optional", "TRUE/FALSE, defaults FALSE."),
    ("is_first_edition", "Optional", "TRUE/FALSE, defaults FALSE."),
    ("image_url", "Optional", "Stock photo URL, if you have one handy."),
    ("condition", "Required", "Near Mint, Lightly Played, Moderately Played, Heavily Played, Damaged."),
    ("quantity", "Required", "How many copies at this condition/price."),
    ("price", "Required", "Price paid per card, USD."),
    ("purchase_date", "Optional", "YYYY-MM-DD. Leave blank to default to today on import."),
    ("source", "Optional", "tcgplayer, ebay, local_shop, card_show, trade, other."),
    ("reference_id", "Optional", "Order/reference ID, if any."),
    ("notes", "Optional", "Anything else worth recording."),
]

wb = Workbook()

# --- Sheet 1: Cards ---
ws = wb.active
ws.title = "Cards"
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")

for col, name in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for r, row in enumerate(EXAMPLES, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)

ws.freeze_panes = "A2"
for col in range(1, len(HEADERS) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 16

# --- Sheet 2: Field Guide ---
guide = wb.create_sheet("Field Guide")
for r, row in enumerate(FIELD_GUIDE, start=1):
    for c, val in enumerate(row, start=1):
        cell = guide.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = header_font
            cell.fill = header_fill
guide.column_dimensions["A"].width = 18
guide.column_dimensions["B"].width = 18
guide.column_dimensions["C"].width = 90
guide.freeze_panes = "A2"

wb.save("docs/plans/card_import_template.xlsx")
print("Wrote docs/plans/card_import_template.xlsx")
