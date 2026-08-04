"""
importer/excel_staging.py
Imports a filled-out spreadsheet (see docs/plans/card_import_template.xlsx
for the expected columns) into staging.

Resolution order per row, same "why" as every other importer in this repo
(the API is just a way to resolve to a card_master row — prefer our own
data when we already have it):
  1. Local match — set + printed card number against our own card_master.
     No network call.
  2. PokemonTCG API fallback for anything not found locally (fired in
     parallel across all such rows, since each is an independent HTTP
     call — same ThreadPoolExecutor pattern as
     importer/market_price_refresh.py). A single API match creates the
     card_master row from real API data, same as every other importer.
  3. Manual fallback — if the API doesn't have it either, create the card
     directly from the row's own columns. This is the only path with no
     external data backing it, which is why it's last, not first.

A row matching more than one existing card (locally or via the API) is
left "ambiguous" for --review to resolve by hand.
"""

from datetime import datetime, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook

from db.connection import (
    get_game_id, find_set_by_name, get_or_create_set,
    find_card_by_number_set, find_card_by_external_id,
    insert_card_master, insert_card_attributes
)
from db.staging import create_batch_id, insert_staging_row, update_staging_row

REQUIRED_COLUMNS = ["card_name", "set_name", "card_number",
                    "condition", "quantity", "price"]
VALID_CONDITIONS = {"Near Mint", "Lightly Played", "Moderately Played",
                    "Heavily Played", "Damaged"}
TRUE_STRINGS = {"true", "yes", "y", "1"}
API_WORKERS = 15  # matches market_price_refresh.py's precedent


def import_from_excel(path: str, dry_run: bool = False, verbose: bool = True) -> dict:
    p = Path(path).expanduser()
    if not p.exists():
        print(f"File not found: {path}")
        return {}

    wb = load_workbook(p, data_only=True)
    ws = wb["Cards"] if "Cards" in wb.sheetnames else wb.worksheets[0]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [(c.value or "").strip() if isinstance(c.value, str) else c.value
               for c in header_row]
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        print(f"Missing required column(s) in '{ws.title}': {', '.join(missing_cols)}")
        return {}

    batch_id = create_batch_id()
    _print(verbose, "\n=== Excel Import ===")
    _print(verbose, f"Batch: {batch_id}")
    _print(verbose, f"File:  {p.name}\n")

    game_id = get_game_id("Pokemon")

    # ── Pass 1: parse + validate every row, resolve what we can locally ──────
    entries = []       # every row that passed basic validation, in order
    needs_api = []      # subset of the above that needs an API lookup
    resolved = {}       # row_idx -> (card_id, match_status, options, created, via)
    skipped = 0

    for row_idx, cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row = dict(zip(headers, [c.value for c in cells]))
        if not row.get("card_name"):
            continue  # blank row (e.g. trailing rows in the sheet)

        missing = [c for c in REQUIRED_COLUMNS if row.get(c) in (None, "")]
        if missing:
            _print(verbose, f"  x Row {row_idx}: missing {', '.join(missing)} — skipped")
            skipped += 1
            continue

        condition = str(row["condition"]).strip()
        if condition not in VALID_CONDITIONS:
            _print(verbose, f"  x Row {row_idx}: unrecognized condition "
                            f"'{condition}' — skipped")
            skipped += 1
            continue

        try:
            quantity = int(row["quantity"])
            price = float(row["price"])
        except (TypeError, ValueError):
            _print(verbose, f"  x Row {row_idx}: quantity/price not numeric — skipped")
            skipped += 1
            continue

        entry = {"row_idx": row_idx, "row": row, "condition": condition,
                 "quantity": quantity, "price": price}
        entries.append(entry)

        local = _try_local_match(row)
        if local is None:
            needs_api.append(entry)
        else:
            resolved[row_idx] = local

    # ── Pass 2: API fallback for everything not found locally, in parallel ──
    if needs_api:
        _print(verbose, f"  Checking PokemonTCG API for {len(needs_api)} "
                        f"card(s) not found locally (parallel, {API_WORKERS} "
                        f"workers)...")
        resolved.update(_resolve_via_api(needs_api, game_id, dry_run))

    # ── Pass 3: write to staging + print, in original row order ─────────────
    totals = {"staged": 0, "matched": 0, "ambiguous": 0, "skipped": skipped,
              "created_api": 0, "created_manual": 0}

    for entry in entries:
        row_idx, row = entry["row_idx"], entry["row"]
        card_id, match_status, options, created, via = resolved[row_idx]

        if card_id is None and match_status == "not_found":
            _print(verbose, f"  x Row {row_idx}: {options[0]['reason']} — skipped")
            totals["skipped"] += 1
            continue

        purchase_date = _parse_date(row.get("purchase_date"))
        order_number = str(row.get("reference_id") or "").strip() or None
        source = str(row.get("source") or "other").strip()

        if not dry_run:
            staging_id = insert_staging_row(
                batch_id=batch_id,
                order_number=order_number,
                order_date=purchase_date,
                card_name=str(row["card_name"]).strip(),
                set_name=str(row["set_name"]).strip(),
                condition=entry["condition"],
                quantity=entry["quantity"],
                price=entry["price"],
                source=source,
                card_id=card_id,
                match_status=match_status,
                match_options=options,
                foil_type=row.get("foil_type"),
                foil_pattern=row.get("foil_pattern"),
                texture=row.get("texture"),
                material=row.get("material"),
                size=row.get("size"),
                stamp_type=row.get("stamp_type"),
                source_type=row.get("source_type"),
                is_shiny=_to_bool(row.get("is_shiny")),
            )
            if match_status == "matched":
                update_staging_row(staging_id, status="approved")

        totals["staged"] += 1
        if match_status == "matched":
            totals["matched"] += 1
            if created and via == "api":
                totals["created_api"] += 1
            elif created and via == "manual":
                totals["created_manual"] += 1
        elif match_status == "ambiguous":
            totals["ambiguous"] += 1

        icon = "+" if created else ("v" if match_status == "matched" else "?")
        note = f" (new card, {via})" if created else (" (local match)" if via == "local" else "")
        _print(verbose,
               f"  {icon} Row {row_idx}: {entry['quantity']}x {row['card_name']} "
               f"#{row['card_number']} [{entry['condition']}] ${entry['price']:.2f}{note}")

    created_total = totals["created_api"] + totals["created_manual"]
    _print(verbose, f"\n{'[DRY RUN] ' if dry_run else ''}Import complete.")
    _print(verbose, f"  Staged:    {totals['staged']} row(s)")
    _print(verbose, f"  Matched:   {totals['matched']} ({created_total} newly created — "
                    f"{totals['created_api']} from API, {totals['created_manual']} manual)")
    _print(verbose, f"  Ambiguous: {totals['ambiguous']} need you to pick the right card")
    _print(verbose, f"  Skipped:   {totals['skipped']} (bad/incomplete rows)")

    if not dry_run and totals["staged"] > 0:
        _print(verbose, "\nNext steps:")
        if totals["ambiguous"]:
            _print(verbose, "  python3 main.py --review   -> resolve ambiguous matches")
        _print(verbose, "  python3 main.py --approve  -> push approved rows to inventory")

    return {**totals, "batch_id": batch_id}


def _try_local_match(row: dict):
    """
    Attempt to resolve purely from our own catalog, no network call.
    Returns a resolved (card_id, match_status, options, created, via) tuple
    if this row is fully decided locally, or None if it needs the API
    fallback (unknown set, or no local card_number match).
    """
    set_name = str(row["set_name"]).strip()
    card_number = str(row["card_number"]).strip()

    existing_set = find_set_by_name(set_name)
    if not existing_set:
        return None  # let the API try — it can resolve/create the set itself

    candidates = find_card_by_number_set(str(existing_set["id"]), card_number)

    if len(candidates) == 1:
        return str(candidates[0]["id"]), "matched", None, False, "local"

    if len(candidates) > 1:
        options = [{"card_id": str(c["id"]), "name": c["name"],
                    "card_number": c["card_number"], "rarity": c.get("rarity")}
                   for c in candidates]
        return None, "ambiguous", options, False, "local"

    return None  # no local match — needs the API fallback


def _resolve_via_api(needs_api: list, game_id: str, dry_run: bool) -> dict:
    """Run the PokemonTCG API search for every row that needs it in
    parallel (each is an independent HTTP call), then finalize each result
    (writes) sequentially — avoids any concurrent-insert race on
    get_or_create_set/insert_card_master."""
    from utils.pokemon_api import search_cards

    def _search(entry):
        row = entry["row"]
        results = search_cards(
            name=str(row["card_name"]).strip(),
            set_name=str(row["set_name"]).strip(),
            card_number=str(row["card_number"]).strip(),
        )
        return entry["row_idx"], results

    api_results_by_row = {}
    with ThreadPoolExecutor(max_workers=API_WORKERS) as pool:
        futures = [pool.submit(_search, e) for e in needs_api]
        for f in as_completed(futures):
            row_idx, results = f.result()
            api_results_by_row[row_idx] = results

    resolved = {}
    for entry in needs_api:
        row_idx = entry["row_idx"]
        resolved[row_idx] = _finalize_api_match(
            entry["row"], api_results_by_row[row_idx], game_id, dry_run
        )
    return resolved


def _finalize_api_match(row: dict, api_results: list, game_id: str, dry_run: bool):
    if len(api_results) > 1:
        options = [{"id": c["id"], "name": c["name"],
                    "card_number": c.get("number"),
                    "set": c["set"]["name"],
                    "variant": ", ".join(c.get("subtypes", []))}
                   for c in api_results]
        return None, "ambiguous", options, False, "api"

    if len(api_results) == 1:
        api_card = api_results[0]
        if dry_run:
            return "dry-run-id", "matched", None, True, "api"

        existing = find_card_by_external_id(api_card["id"])
        if existing:
            return str(existing["id"]), "matched", None, False, "api"

        from utils.pokemon_api import parse_card_master_fields, parse_card_attribute_fields
        fields = parse_card_master_fields(api_card)
        attr_fields = parse_card_attribute_fields(api_card)
        set_id = get_or_create_set(
            game_id=game_id, name=fields["set_name"], set_code=fields["set_code"],
            series=fields.get("series"), release_year=fields.get("release_year"),
            total_cards=fields.get("total_cards"),
        )
        card_id = insert_card_master(
            set_id=set_id, name=fields["name"], card_number=fields["card_number"],
            rarity=fields.get("rarity"), is_promo=fields.get("is_promo", False),
            is_first_edition=fields.get("is_first_edition", False),
            image_url=fields.get("image_url"), external_id=fields["external_id"],
        )
        insert_card_attributes(card_id, **attr_fields)
        return card_id, "matched", None, True, "api"

    # Zero API results — final fallback: create directly from the
    # spreadsheet's own data. Only path with no external data backing it.
    return _manual_create(row, game_id, dry_run)


def _manual_create(row: dict, game_id: str, dry_run: bool):
    set_name = str(row["set_name"]).strip()
    set_code = str(row.get("set_code") or "").strip() or None
    card_name = str(row["card_name"]).strip()
    card_number = str(row["card_number"]).strip()

    existing_set = find_set_by_name(set_name)
    if existing_set:
        set_id = str(existing_set["id"])
    elif set_code:
        if dry_run:
            return "dry-run-id", "matched", None, True, "manual"
        set_id = get_or_create_set(game_id=game_id, name=set_name, set_code=set_code)
    else:
        return None, "not_found", [{
            "reason": f"Unknown set '{set_name}' — set_code required to create it "
                      f"(also not found via the PokemonTCG API)"
        }], False, "manual"

    if dry_run:
        return "dry-run-id", "matched", None, True, "manual"

    card_id = insert_card_master(
        set_id=set_id,
        name=card_name,
        card_number=card_number,
        rarity=(str(row["rarity"]).strip() if row.get("rarity") else None),
        is_promo=_to_bool(row.get("is_promo")),
        is_first_edition=_to_bool(row.get("is_first_edition")),
        image_url=(str(row["image_url"]).strip() if row.get("image_url") else None),
    )
    return card_id, "matched", None, True, "manual"


def _to_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    return str(val).strip().lower() in TRUE_STRINGS


def _parse_date(val):
    if val is None or val == "":
        return datetime.now(timezone.utc)
    if isinstance(val, datetime):
        return val if val.tzinfo else val.replace(tzinfo=timezone.utc)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return datetime.now(timezone.utc)


def _print(verbose: bool, msg: str):
    if verbose:
        print(msg)
